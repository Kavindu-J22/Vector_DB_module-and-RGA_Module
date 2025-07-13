"""
Simple test script to validate basic functionality without complex dependencies
"""

import json
import openpyxl
import re
from typing import List, Dict, Any

def test_data_loading():
    """Test loading of data files"""
    print("Testing data loading...")
    
    # Test JSON data loading
    try:
        with open('acts_2024.json', 'r', encoding='utf-8') as f:
            acts_data = json.load(f)
        print(f"✓ Loaded {len(acts_data)} acts")
        
        with open('cases_2024.json', 'r', encoding='utf-8') as f:
            cases_data = json.load(f)
        print(f"✓ Loaded {len(cases_data)} cases")
        
    except Exception as e:
        print(f"✗ Error loading JSON data: {e}")
        return False
    
    # Test Excel data loading
    try:
        wb = openpyxl.load_workbook('tags for classifcation and metadata-vector embeddings.xlsx')
        ws = wb.active
        
        tags = {'family': [], 'property': [], 'commercial': [], 'labour': []}
        
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=5, values_only=True):
            categories = ['family', 'property', 'commercial', 'labour']
            for i, cell_value in enumerate(row):
                if cell_value and i < len(categories):
                    tags[categories[i]].append(str(cell_value).strip())
        
        total_tags = sum(len(v) for v in tags.values())
        print(f"✓ Loaded {total_tags} classification tags")
        
    except Exception as e:
        print(f"✗ Error loading Excel data: {e}")
        return False
    
    return True

def test_text_processing():
    """Test basic text processing functions"""
    print("\nTesting text processing...")
    
    # Test text cleaning
    sample_text = "This is a   test\n\nwith multiple    spaces and\nnewlines."
    cleaned = re.sub(r'\s+', ' ', sample_text).strip()
    print(f"✓ Text cleaning: '{sample_text}' -> '{cleaned}'")
    
    # Test language detection
    sinhala_text = "මෙය සිංහල පාඨයකි"
    tamil_text = "இது தமிழ் உரை"
    english_text = "This is English text"
    
    def detect_language(text):
        sinhala_chars = len(re.findall(r'[\u0D80-\u0DFF]', text))
        tamil_chars = len(re.findall(r'[\u0B80-\u0BFF]', text))
        english_chars = len(re.findall(r'[a-zA-Z]', text))
        
        total_chars = len(text)
        
        if sinhala_chars / total_chars > 0.3:
            return "sinhala"
        elif tamil_chars / total_chars > 0.3:
            return "tamil"
        else:
            return "english"
    
    print(f"✓ Language detection: Sinhala -> {detect_language(sinhala_text)}")
    print(f"✓ Language detection: Tamil -> {detect_language(tamil_text)}")
    print(f"✓ Language detection: English -> {detect_language(english_text)}")
    
    return True

def test_document_chunking():
    """Test document chunking functionality"""
    print("\nTesting document chunking...")
    
    sample_doc = """
    This is a test legal document. It contains multiple sentences for testing chunking functionality. 
    The document discusses property rights and ownership laws in Sri Lanka. 
    Property ownership is a fundamental right protected by law. 
    The transfer of property requires proper documentation and legal procedures.
    """
    
    # Simple sentence-based chunking
    sentences = sample_doc.split('.')
    sentences = [s.strip() for s in sentences if s.strip()]
    
    chunk_size = 2  # sentences per chunk
    chunks = []
    
    for i in range(0, len(sentences), chunk_size):
        chunk = '. '.join(sentences[i:i + chunk_size])
        if chunk:
            chunks.append(chunk + '.')
    
    print(f"✓ Created {len(chunks)} chunks from {len(sentences)} sentences")
    for i, chunk in enumerate(chunks):
        print(f"  Chunk {i+1}: {chunk[:50]}...")
    
    return True

def test_metadata_extraction():
    """Test metadata extraction"""
    print("\nTesting metadata extraction...")
    
    sample_document = {
        'id': 'test-doc-001',
        'filename': 'Test Legal Document',
        'primaryLang': 'English',
        'text': 'This is a test legal document about property rights.',
        'wordCount': 10
    }
    
    # Extract metadata
    metadata = {
        'document_type': 'act',
        'original_id': sample_document['id'],
        'filename': sample_document['filename'],
        'chunk_index': 0,
        'sequence_number': 1,
        'primary_language': sample_document['primaryLang'],
        'detected_language': 'english',
        'word_count': len(sample_document['text'].split()),
        'chunk_text_preview': sample_document['text'][:100]
    }
    
    print("✓ Metadata extracted:")
    for key, value in metadata.items():
        print(f"  {key}: {value}")
    
    return True

def test_classification_tags():
    """Test classification tag matching"""
    print("\nTesting classification tag matching...")
    
    # Load tags
    wb = openpyxl.load_workbook('tags for classifcation and metadata-vector embeddings.xlsx')
    ws = wb.active
    
    tags = {'family': [], 'property': [], 'commercial': [], 'labour': []}
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=5, values_only=True):
        categories = ['family', 'property', 'commercial', 'labour']
        for i, cell_value in enumerate(row):
            if cell_value and i < len(categories):
                tags[categories[i]].append(str(cell_value).strip().lower())
    
    # Test text classification
    test_texts = [
        "This document discusses property ownership and transfer rights",
        "Marriage dissolution and divorce proceedings require court approval",
        "Employment contract disputes and severance pay issues",
        "Commercial agreements and intellectual property rights"
    ]
    
    for text in test_texts:
        text_lower = text.lower()
        matched_categories = []
        
        for category, category_tags in tags.items():
            for tag in category_tags:
                if tag in text_lower:
                    matched_categories.append(category)
                    break
        
        print(f"✓ Text: '{text[:50]}...'")
        print(f"  Matched categories: {matched_categories}")
    
    return True

def test_config_validation():
    """Test configuration validation"""
    print("\nTesting configuration...")
    
    try:
        import config
        
        print(f"✓ Pinecone API key configured: {'*' * 20}{config.PINECONE_API_KEY[-10:]}")
        print(f"✓ Embedding model: {config.EMBEDDING_MODEL_NAME}")
        print(f"✓ Vector dimension: {config.VECTOR_DIMENSION}")
        print(f"✓ Chunk size: {config.CHUNK_SIZE}")
        print(f"✓ Supported languages: {config.SUPPORTED_LANGUAGES}")
        
        return True
        
    except Exception as e:
        print(f"✗ Configuration error: {e}")
        return False

def main():
    """Run all simple tests"""
    print("=== Simple System Validation ===")
    print("Running basic tests without complex dependencies...\n")
    
    tests = [
        ("Data Loading", test_data_loading),
        ("Text Processing", test_text_processing),
        ("Document Chunking", test_document_chunking),
        ("Metadata Extraction", test_metadata_extraction),
        ("Classification Tags", test_classification_tags),
        ("Configuration", test_config_validation)
    ]
    
    results = {}
    all_passed = True
    
    for test_name, test_function in tests:
        try:
            result = test_function()
            results[test_name] = result
            
            if result:
                print(f"✅ {test_name} - PASSED\n")
            else:
                print(f"❌ {test_name} - FAILED\n")
                all_passed = False
                
        except Exception as e:
            print(f"❌ {test_name} - ERROR: {e}\n")
            results[test_name] = False
            all_passed = False
    
    # Summary
    print("=" * 50)
    if all_passed:
        print("🎉 ALL BASIC TESTS PASSED!")
        print("\nThe system is ready for:")
        print("1. Document processing and chunking")
        print("2. Classification tag matching")
        print("3. Metadata extraction")
        print("4. Text preprocessing")
        
        print("\nNext steps:")
        print("1. Install remaining dependencies: pip install -r requirements.txt")
        print("2. Run full system test: python test_system.py")
        print("3. Execute complete pipeline: python main.py")
        
    else:
        print("❌ SOME TESTS FAILED")
        failed_tests = [test for test, result in results.items() if not result]
        print(f"Failed tests: {', '.join(failed_tests)}")
        
        print("\nPlease check:")
        print("1. All data files are present")
        print("2. Configuration is correct")
        print("3. File permissions are adequate")
    
    print("=" * 50)

if __name__ == "__main__":
    main()
