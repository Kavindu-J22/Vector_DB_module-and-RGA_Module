import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('tags for classifcation and metadata-vector embeddings.xlsx')
ws = wb.active

print(f'Sheet name: {ws.title}')
print(f'Max row: {ws.max_row}')
print(f'Max col: {ws.max_column}')

print('\nFirst 20 rows:')
for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
    print(row)
