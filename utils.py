"""
Utility functions for the Vector Database and Embedding Module
"""

import json
import re
import logging
import openpyxl
from typing import List, Dict, Any, Tuple
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
import config

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt')

try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords')

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def load_json_data(file_path: str) -> List[Dict[str, Any]]:
    """Load JSON data from file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Loaded {len(data)} documents from {file_path}")
        return data
    except Exception as e:
        logger.error(f"Error loading {file_path}: {str(e)}")
        return []

def load_classification_tags() -> Dict[str, List[str]]:
    """Load classification tags from Excel file"""
    try:
        wb = openpyxl.load_workbook(config.TAGS_EXCEL_PATH)
        ws = wb.active
        
        tags = {
            'family': [],
            'property': [],
            'commercial': [],
            'labour': []
        }
        
        # Extract tags from Excel (starting from row 4, columns B-E)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=5, values_only=True):
            categories = ['family', 'property', 'commercial', 'labour']
            for i, cell_value in enumerate(row):
                if cell_value and i < len(categories):
                    tags[categories[i]].append(str(cell_value).strip())
        
        # Remove empty strings and None values
        for category in tags:
            tags[category] = [tag for tag in tags[category] if tag and tag.strip()]
        
        logger.info(f"Loaded classification tags: {sum(len(v) for v in tags.values())} total tags")
        return tags
        
    except Exception as e:
        logger.error(f"Error loading classification tags: {str(e)}")
        return {}

def clean_text(text: str) -> str:
    """Clean and preprocess text"""
    if not text:
        return ""
    
    # Remove extra whitespace and newlines
    text = re.sub(r'\s+', ' ', text)
    
    # Remove special characters but keep basic punctuation
    text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
    
    # Remove multiple consecutive punctuation
    text = re.sub(r'([.,;:!?]){2,}', r'\1', text)
    
    return text.strip()

def detect_language(text: str) -> str:
    """Simple language detection for Sinhala, Tamil, and English"""
    if not text:
        return "english"
    
    # Count characters in different scripts
    sinhala_chars = len(re.findall(r'[\u0D80-\u0DFF]', text))
    tamil_chars = len(re.findall(r'[\u0B80-\u0BFF]', text))
    english_chars = len(re.findall(r'[a-zA-Z]', text))
    
    total_chars = len(text)
    
    if sinhala_chars / total_chars > 0.3:
        return "sinhala"
    elif tamil_chars / total_chars > 0.3:
        return "tamil"
    else:
        return "english"

def chunk_text(text: str, chunk_size: int = config.CHUNK_SIZE, overlap: int = config.CHUNK_OVERLAP) -> List[str]:
    """Split text into overlapping chunks"""
    if not text:
        return []
    
    # Tokenize into sentences
    sentences = sent_tokenize(text)
    
    chunks = []
    current_chunk = ""
    current_length = 0
    
    for sentence in sentences:
        sentence_length = len(word_tokenize(sentence))
        
        if current_length + sentence_length <= chunk_size:
            current_chunk += " " + sentence if current_chunk else sentence
            current_length += sentence_length
        else:
            if current_chunk:
                chunks.append(current_chunk.strip())
            
            # Start new chunk with overlap
            if overlap > 0 and chunks:
                # Take last few words from previous chunk for overlap
                words = word_tokenize(current_chunk)
                overlap_words = words[-overlap:] if len(words) > overlap else words
                current_chunk = " ".join(overlap_words) + " " + sentence
                current_length = len(overlap_words) + sentence_length
            else:
                current_chunk = sentence
                current_length = sentence_length
    
    # Add the last chunk
    if current_chunk:
        chunks.append(current_chunk.strip())
    
    return chunks

def create_document_id(doc_type: str, original_id: str, chunk_index: int) -> str:
    """Create unique document ID for chunks"""
    return f"{doc_type}_{original_id}_{chunk_index:04d}"

def extract_metadata(document: Dict[str, Any], chunk_index: int, chunk_text: str) -> Dict[str, Any]:
    """Extract metadata for document chunk"""
    language = detect_language(chunk_text)
    
    metadata = {
        'document_type': 'act' if 'Act' in document.get('filename', '') else 'case',
        'original_id': document.get('id', ''),
        'filename': document.get('filename', ''),
        'chunk_index': chunk_index,
        'sequence_number': chunk_index + 1,
        'primary_language': document.get('primaryLang', 'english'),
        'detected_language': language,
        'word_count': len(word_tokenize(chunk_text)),
        'chunk_text_preview': chunk_text[:100] + "..." if len(chunk_text) > 100 else chunk_text
    }
    
    return metadata

def save_processed_data(data: List[Dict[str, Any]], filename: str):
    """Save processed data to JSON file"""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Saved {len(data)} processed documents to {filename}")
    except Exception as e:
        logger.error(f"Error saving processed data: {str(e)}")

def load_processed_data(filename: str) -> List[Dict[str, Any]]:
    """Load processed data from JSON file"""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"Loaded {len(data)} processed documents from {filename}")
        return data
    except Exception as e:
        logger.error(f"Error loading processed data: {str(e)}")
        return []
