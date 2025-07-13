"""
Simplified Real Data Upload - Bypasses numpy compatibility issues
Processes real legal documents with chunking, sequence numbers, and embeddings
"""

from pinecone import Pinecone
import json
import time
import re
import openpyxl
from sentence_transformers import SentenceTransformer
from typing import List, Dict, Any
import config

def load_classification_tags() -> Dict[str, List[str]]:
    """Load classification tags from Excel file"""
    try:
        wb = openpyxl.load_workbook('tags for classifcation and metadata-vector embeddings.xlsx')
        ws = wb.active
        
        tags = {'family': [], 'property': [], 'commercial': [], 'labour': []}
        
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=5, values_only=True):
            categories = ['family', 'property', 'commercial', 'labour']
            for i, cell_value in enumerate(row):
                if cell_value and i < len(categories):
                    tags[categories[i]].append(str(cell_value).strip().lower())
        
        return tags
    except Exception as e:
        print(f"Error loading tags: {e}")
        return {}

def clean_text(text: str) -> str:
    """Clean and preprocess text"""
    if not text:
        return ""
    
    # Remove extra whitespace and newlines
    text = re.sub(r'\s+', ' ', text)
    
    # Remove special characters but keep basic punctuation
    text = re.sub(r'[^\w\s.,;:!?()-]', '', text)
    
    return text.strip()

def detect_language(text: str) -> str:
    """Simple language detection"""
    if not text:
        return "english"
    
    sinhala_chars = len(re.findall(r'[\u0D80-\u0DFF]', text))
    tamil_chars = len(re.findall(r'[\u0B80-\u0BFF]', text))
    english_chars = len(re.findall(r'[a-zA-Z]', text))
    
    total_chars = len(text)
    
    if sinhala_chars / total_chars > 0.3:
        return "sinhala"
    elif tamil_chars / total_chars > 0.3:
        return "tamil"
    else:
        return "english"

def chunk_text(text: str, chunk_size: int = 400) -> List[str]:
    """Split text into chunks by words"""
    if not text:
        return []
    
    words = text.split()
    chunks = []
    
    for i in range(0, len(words), chunk_size):
        chunk = ' '.join(words[i:i + chunk_size])
        if chunk.strip():
            chunks.append(chunk.strip())
    
    return chunks

def classify_text(text: str, classification_tags: Dict[str, List[str]]) -> Dict[str, Any]:
    """Simple keyword-based classification"""
    text_lower = text.lower()
    matched_categories = []
    matched_tags = []
    
    for category, tags in classification_tags.items():
        category_matches = 0
        for tag in tags:
            if tag in text_lower:
                matched_tags.append(tag)
                category_matches += 1
        
        if category_matches > 0:
            matched_categories.append(category)
    
    # Determine primary category
    if matched_categories:
        primary_category = matched_categories[0].capitalize()
    else:
        primary_category = "General"
    
    return {
        'primary_category': primary_category,
        'matched_categories': matched_categories,
        'classification_tags': matched_tags[:5]  # Limit to 5 tags
    }

def process_document(doc: Dict[str, Any], doc_type: str, classification_tags: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    """Process a single document into chunks"""
    text = doc.get('text', '')
    if not text:
        return []
    
    # Clean text
    cleaned_text = clean_text(text)
    
    # Split into chunks
    chunks = chunk_text(cleaned_text)
    
    processed_chunks = []
    for i, chunk_text in enumerate(chunks):
        if not chunk_text.strip():
            continue
        
        # Classify chunk
        classification = classify_text(chunk_text, classification_tags)
        
        # Create metadata
        metadata = {
            'document_type': doc_type,
            'original_id': doc.get('id', ''),
            'filename': doc.get('filename', '')[:100],  # Limit length
            'chunk_index': i,
            'sequence_number': i + 1,
            'primary_language': doc.get('primaryLang', 'English'),
            'detected_language': detect_language(chunk_text),
            'word_count': len(chunk_text.split()),
            'text_preview': chunk_text[:200],  # First 200 chars
            'primary_category': classification['primary_category'],
            'classification_tags': ','.join(classification['classification_tags'])
        }
        
        # Create chunk ID
        chunk_id = f"{doc_type}_{doc.get('id', 'unknown')}_{i:04d}"
        
        processed_chunk = {
            'id': chunk_id,
            'text': chunk_text,
            'metadata': metadata
        }
        
        processed_chunks.append(processed_chunk)
    
    return processed_chunks

def generate_embeddings(chunks: List[Dict[str, Any]], model_name: str = "sentence-transformers/all-MiniLM-L6-v2") -> List[Dict[str, Any]]:
    """Generate embeddings for chunks"""
    print(f"🧠 Loading embedding model: {model_name}")
    
    try:
        model = SentenceTransformer(model_name)
        print(f"✅ Model loaded successfully")
        
        # Extract texts
        texts = [chunk['text'] for chunk in chunks]
        
        print(f"🔄 Generating embeddings for {len(texts)} chunks...")
        
        # Generate embeddings in batches
        batch_size = 32
        all_embeddings = []
        
        for i in range(0, len(texts), batch_size):
            batch_texts = texts[i:i + batch_size]
            batch_embeddings = model.encode(batch_texts, convert_to_numpy=True, show_progress_bar=False)
            all_embeddings.extend(batch_embeddings)
            
            if (i // batch_size + 1) % 10 == 0:
                print(f"   Processed {i + len(batch_texts)} / {len(texts)} chunks")
        
        # Add embeddings to chunks
        embedded_chunks = []
        for chunk, embedding in zip(chunks, all_embeddings):
            chunk['embedding'] = embedding.tolist()
            chunk['metadata']['embedding_model'] = model_name
            chunk['metadata']['embedding_dimension'] = len(embedding)
            embedded_chunks.append(chunk)
        
        print(f"✅ Generated embeddings for {len(embedded_chunks)} chunks")
        return embedded_chunks
        
    except Exception as e:
        print(f"❌ Error generating embeddings: {e}")
        return chunks

def upload_to_pinecone(embedded_chunks: List[Dict[str, Any]]) -> bool:
    """Upload embedded chunks to Pinecone"""
    print(f"📤 Uploading {len(embedded_chunks)} chunks to Pinecone...")
    
    try:
        pc = Pinecone(api_key=config.PINECONE_API_KEY)
        index = pc.Index(config.PINECONE_INDEX_NAME)
        
        # Prepare vectors
        vectors = []
        for chunk in embedded_chunks:
            if 'embedding' not in chunk:
                continue
            
            # Prepare metadata for Pinecone
            metadata = {}
            for key, value in chunk['metadata'].items():
                if isinstance(value, (str, int, float, bool)):
                    if isinstance(value, str) and len(value) > 1000:
                        metadata[key] = value[:1000]  # Truncate long strings
                    else:
                        metadata[key] = value
            
            vector = {
                'id': chunk['id'],
                'values': chunk['embedding'],
                'metadata': metadata
            }
            vectors.append(vector)
        
        print(f"📋 Prepared {len(vectors)} vectors for upload")
        
        # Upload in batches
        batch_size = 100
        for i in range(0, len(vectors), batch_size):
            batch = vectors[i:i + batch_size]
            index.upsert(vectors=batch)
            
            print(f"   Uploaded batch {i//batch_size + 1}/{(len(vectors)-1)//batch_size + 1}")
            time.sleep(1)  # Small delay between batches
        
        print(f"✅ Successfully uploaded {len(vectors)} vectors")
        return True
        
    except Exception as e:
        print(f"❌ Error uploading to Pinecone: {e}")
        return False

def main():
    """Main function to process and upload real legal documents"""
    print("=" * 70)
    print("📚 REAL LEGAL DOCUMENTS PROCESSOR (Simplified)")
    print("=" * 70)
    
    # Load data
    print("📋 Loading legal documents...")
    
    try:
        with open('acts_2024.json', 'r', encoding='utf-8') as f:
            acts_data = json.load(f)
        
        with open('cases_2024.json', 'r', encoding='utf-8') as f:
            cases_data = json.load(f)
        
        print(f"✅ Loaded {len(acts_data)} acts and {len(cases_data)} cases")
        
    except Exception as e:
        print(f"❌ Error loading data: {e}")
        return
    
    # Load classification tags
    print("🏷️ Loading classification tags...")
    classification_tags = load_classification_tags()
    total_tags = sum(len(tags) for tags in classification_tags.values())
    print(f"✅ Loaded {total_tags} classification tags")
    
    # Clear existing data
    print("\n🧹 Clearing existing data from Pinecone...")
    try:
        pc = Pinecone(api_key=config.PINECONE_API_KEY)
        index = pc.Index(config.PINECONE_INDEX_NAME)
        index.delete(delete_all=True)
        print("✅ Existing data cleared")
        time.sleep(5)
    except Exception as e:
        print(f"❌ Error clearing data: {e}")
        return
    
    # Process documents
    print("\n📄 Processing documents into chunks...")
    
    all_chunks = []
    
    # Process Acts
    print("   Processing Acts...")
    for doc in acts_data:
        chunks = process_document(doc, 'act', classification_tags)
        all_chunks.extend(chunks)
    
    # Process Cases  
    print("   Processing Cases...")
    for doc in cases_data:
        chunks = process_document(doc, 'case', classification_tags)
        all_chunks.extend(chunks)
    
    print(f"✅ Generated {len(all_chunks)} total chunks")
    
    # Generate embeddings
    print("\n🧠 Generating embeddings...")
    embedded_chunks = generate_embeddings(all_chunks)
    
    # Upload to Pinecone
    print("\n📤 Uploading to Pinecone...")
    if upload_to_pinecone(embedded_chunks):
        print("\n🎉 SUCCESS!")
        print(f"✅ Processed and uploaded {len(embedded_chunks)} real legal document chunks")
        
        # Final verification
        print("\n🔍 Verifying upload...")
        time.sleep(10)
        
        try:
            stats = index.describe_index_stats()
            print(f"📊 Final Pinecone statistics:")
            print(f"   Total vectors: {stats.get('total_vector_count', 0)}")
            print(f"   Dimension: {stats.get('dimension', 0)}")
            
            # Test search
            print("\n🔍 Testing search...")
            query_embedding = embedded_chunks[0]['embedding']
            results = index.query(vector=query_embedding, top_k=3, include_metadata=True)
            
            print(f"📋 Search test returned {len(results.get('matches', []))} results")
            for i, match in enumerate(results.get('matches', [])):
                print(f"   {i+1}. {match['metadata'].get('document_type', 'unknown')} - {match['metadata'].get('primary_category', 'unknown')}")
            
        except Exception as e:
            print(f"⚠️ Verification error: {e}")
        
        print(f"\n🎯 Your Pinecone dashboard now shows REAL legal document data!")
        print(f"   📄 {len(acts_data)} Acts processed into chunks")
        print(f"   📄 {len(cases_data)} Cases processed into chunks") 
        print(f"   🔢 Each chunk has sequence numbers")
        print(f"   🧠 Real embeddings generated")
        print(f"   🏷️ Legal classification applied")
        
    else:
        print("\n❌ Upload failed")

if __name__ == "__main__":
    main()
